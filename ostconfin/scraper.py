"""
웹 데이터 수집 → 엑셀 저장 스크립트

사내 웹 시스템에 로그인 후 테이블 데이터를 수집하여 지정된 엑셀 파일에 저장.

실행 전 필수 설정:
  - config/.env       : WEB_USER_ID, WEB_PASSWORD
  - config/config.ini : URL, 셀렉터, 엑셀 경로 등

실행 방법:
  pip install -r requirements.txt
  playwright install chromium
  python scraper.py
"""
import os
import io
import sys
import configparser
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
from dotenv import load_dotenv
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout


# ── 설정 로드 ─────────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).parent
CONFIG_PATH = BASE_DIR / "config" / "config.ini"
ENV_PATH = BASE_DIR / "config" / ".env"

load_dotenv(dotenv_path=ENV_PATH)

config = configparser.ConfigParser()
config.read(CONFIG_PATH, encoding="utf-8")

USER_ID      = os.getenv("WEB_USER_ID", "")
PASSWORD     = os.getenv("WEB_PASSWORD", "")
LOGIN_URL    = os.getenv("WEB_LOGIN_URL", "")
DATA_URL     = os.getenv("WEB_DATA_URL", "")
OUTPUT_PATH  = Path(os.getenv("EXCEL_OUTPUT_PATH", ""))

missing = [k for k, v in {
    "WEB_USER_ID": USER_ID,
    "WEB_PASSWORD": PASSWORD,
    "WEB_LOGIN_URL": LOGIN_URL,
    "WEB_DATA_URL": DATA_URL,
    "EXCEL_OUTPUT_PATH": str(OUTPUT_PATH),
}.items() if not v]

if missing:
    print(f"[오류] config/.env 에 다음 항목이 설정되지 않았습니다: {', '.join(missing)}")
    sys.exit(1)


# ── 설정값 읽기 (셀렉터/옵션은 config.ini) ────────────────────────────────────

ID_SEL       = config.get("SELECTORS", "id_selector")
PW_SEL       = config.get("SELECTORS", "pw_selector")
BTN_SEL      = config.get("SELECTORS", "login_btn_selector")
TABLE_SEL    = config.get("SELECTORS", "table_selector", fallback="table")
TABLE_IDX    = config.getint("SELECTORS", "table_index", fallback=0)

SHEET_NAME   = config.get("EXCEL", "sheet_name", fallback="Sheet1")
START_CELL   = config.get("EXCEL", "start_cell", fallback="A2")
WRITE_HEADER = config.getboolean("EXCEL", "write_header", fallback=False)

HEADLESS       = config.getboolean("OPTIONS", "headless", fallback=False)
TIMEOUT_SEC    = config.getint("OPTIONS", "timeout", fallback=30)
TIMEOUT_MS     = TIMEOUT_SEC * 1000


# ── 유틸 함수 ─────────────────────────────────────────────────────────────────

def parse_start_cell(cell_str: str) -> tuple[int, int]:
    """'B3' → (col=2, row=3)"""
    cell_str = cell_str.strip().upper()
    col_str = "".join(c for c in cell_str if c.isalpha())
    row_str = "".join(c for c in cell_str if c.isdigit())
    return column_index_from_string(col_str), int(row_str)


# ── 메인 로직 ─────────────────────────────────────────────────────────────────

def run():
    with sync_playwright() as p:
        # Edge(Chromium) 실행
        browser = p.chromium.launch(
            channel="msedge",   # Edge 사용. Edge 없으면 channel 제거 시 기본 Chromium
            headless=HEADLESS,
        )
        context = browser.new_context()
        page = context.new_page()
        page.set_default_timeout(TIMEOUT_MS)

        try:
            # ── STEP 1: 로그인 ────────────────────────────────────────────────
            print(f"[1/3] 로그인 중... ({LOGIN_URL})")
            page.goto(LOGIN_URL)

            # 페이지 JS(RSA 암호화 스크립트 등) 로딩 완료 대기
            page.wait_for_load_state("domcontentloaded")

            page.locator(ID_SEL).fill(USER_ID)
            page.locator(PW_SEL).fill(PASSWORD)
            page.locator(BTN_SEL).click()

            # 로그인 후 페이지 안정화 대기
            page.wait_for_load_state("networkidle")
            print("      로그인 완료")

            # ── STEP 2: 데이터 페이지로 이동 & 테이블 수집 ───────────────────
            print(f"[2/3] 데이터 수집 중... ({DATA_URL})")
            page.goto(DATA_URL)
            page.wait_for_load_state("networkidle")

            # 테이블 요소가 나타날 때까지 대기
            page.locator(TABLE_SEL).first.wait_for(state="visible")

            # 테이블 HTML 추출
            tables = page.locator(TABLE_SEL).all()
            if TABLE_IDX >= len(tables):
                print(f"[오류] table_index={TABLE_IDX} 이지만 테이블이 {len(tables)}개뿐입니다.")
                print("       config.ini의 table_index 또는 table_selector를 확인하세요.")
                sys.exit(1)

            html = tables[TABLE_IDX].evaluate("el => el.outerHTML")
            dfs = pd.read_html(io.StringIO(html))
            df = dfs[0]

            # 숫자 컬럼 정리 (쉼표 제거)
            for col in df.columns:
                cleaned = df[col].astype(str).str.replace(",", "").str.strip()
                try:
                    df[col] = pd.to_numeric(cleaned, errors="raise")
                except (ValueError, TypeError):
                    pass

            print(f"      수집 완료: {len(df)}행 × {len(df.columns)}열")

        except PlaywrightTimeout:
            print("[오류] 페이지 로딩 시간 초과. config.ini의 timeout 값을 늘려보세요.")
            sys.exit(1)
        finally:
            browser.close()

    # ── STEP 3: 엑셀 저장 ─────────────────────────────────────────────────────
    print(f"[3/3] 엑셀 저장 중... ({OUTPUT_PATH})")

    if not OUTPUT_PATH.exists():
        print(f"[오류] 엑셀 파일을 찾을 수 없습니다: {OUTPUT_PATH}")
        print("       config.ini의 output_path를 확인하세요.")
        sys.exit(1)

    wb = openpyxl.load_workbook(OUTPUT_PATH)
    if SHEET_NAME not in wb.sheetnames:
        print(f"[오류] 시트 '{SHEET_NAME}'가 없습니다. 존재하는 시트: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]
    start_col, start_row = parse_start_cell(START_CELL)

    # 기존 데이터 영역 초기화
    for r in range(start_row, ws.max_row + 1):
        for c in range(start_col, start_col + len(df.columns)):
            ws.cell(row=r, column=c).value = None

    # 헤더 쓰기 (옵션)
    row_offset = 0
    if WRITE_HEADER:
        for c_offset, col_name in enumerate(df.columns):
            ws.cell(row=start_row, column=start_col + c_offset).value = col_name
        row_offset = 1

    # 데이터 쓰기
    for r_offset, row_data in enumerate(df.itertuples(index=False)):
        for c_offset, value in enumerate(row_data):
            cell = ws.cell(
                row=start_row + row_offset + r_offset,
                column=start_col + c_offset,
            )
            cell.value = None if pd.isna(value) else value

    wb.save(OUTPUT_PATH)
    print(f"      저장 완료: {len(df)}행 기입 → {OUTPUT_PATH}")
    print("\n완료! 엑셀 파일을 열어 결과를 확인하세요.")


if __name__ == "__main__":
    run()
