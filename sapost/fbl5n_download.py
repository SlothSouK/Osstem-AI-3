"""
FBL5N 채권 미결항목 다운로드 스크립트

사용법:
  python sapost/fbl5n_download.py --keydate 202503

  --keydate  조회 기준 년월 (YYYYMM). 해당 월 말일로 자동 변환됩니다.

동작 순서:
  1. source_dir 의 파일 목록에서 고객계정(파일명 앞 7자리) 수집
  2. 각 고객계정마다 FBL5N 실행
     - 미결항목 / 특별G/L거래 / 임시항목 선택
     - 기준일 = 해당 월 말일
  3. 전기일자 오름차순 정렬 후 엑셀 로컬 저장
  4. raw_dir 에 {계정코드}-{YYYYMM}.xlsx 로 저장
"""
import re
import sys
import time
import shutil
import argparse
import calendar
from copy import copy as _copy_obj
from datetime import date, datetime
from typing import Any
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))


def _load_account_corp_map() -> dict[str, str]:
    """CLAUDE.md 의 법인코드 목록 테이블에서 고객코드 → 법인명 맵 생성."""
    claude_md = Path(__file__).parent.parent / "CLAUDE.md"
    result: dict[str, str] = {}
    if not claude_md.exists():
        return result
    in_table = False
    for line in claude_md.read_text(encoding="utf-8").splitlines():
        # 헤더 행 감지
        if "법인코드" in line and "법인명" in line and "고객코드" in line:
            in_table = True
            continue
        if in_table:
            if not line.strip().startswith("|"):
                break
            parts = [p.strip() for p in line.strip().strip("|").split("|")]
            if len(parts) >= 3 and parts[2].isdigit():
                result[parts[2]] = parts[1]
    return result


ACCOUNT_CORP_MAP: dict[str, str] = _load_account_corp_map()

from sapost.src.utils import get_config, setup_logger

try:
    import win32com.client
except ImportError:
    print("ERROR: pywin32가 설치되어 있지 않습니다. pip install pywin32 를 실행하세요.")
    sys.exit(1)

from dotenv import load_dotenv
import os


# ──────────────────────────────────────────────────────────
# 헬퍼
# ──────────────────────────────────────────────────────────

def month_start(yyyymm: str) -> str:
    """'202603' → '2026.03.01' (SAP 날짜 형식)"""
    year  = int(yyyymm[:4])
    month = int(yyyymm[4:6])
    return f"{year}.{month:02d}.01"


def month_end(yyyymm: str) -> str:
    """'202603' → '2026.03.31' (SAP 날짜 형식)"""
    year  = int(yyyymm[:4])
    month = int(yyyymm[4:6])
    last_day = calendar.monthrange(year, month)[1]
    return f"{year}.{month:02d}.{last_day:02d}"


def find_source_file(source_dir: Path, account: str) -> Path | None:
    """source_dir 에서 파일명이 account(7자리)로 시작하는 엑셀 파일 반환"""
    for f in source_dir.iterdir():
        if f.is_file() and f.stem.startswith(account) and f.suffix in (".xlsx", ".xlsm"):
            return f
    return None


def make_working_copy(source_file: Path, account: str, budat_high: str) -> Path:
    """
    원본 파일을 복사한 작업 복사본 경로 반환.
    파일명: [고객코드] [법인명]법인채권명세서_[최종조회일자YYYYMMDD].xlsx
    budat_high: 'YYYY.MM.DD' 형식
    """
    corp_name = ACCOUNT_CORP_MAP.get(account, account)
    date_str  = budat_high.replace(".", "")   # '20260331'
    new_name  = f"[{account}] {corp_name}법인채권명세서_{date_str}{source_file.suffix}"
    dest      = source_file.parent / new_name
    shutil.copy2(source_file, dest)
    return dest


def _parse_sap_date(val) -> date | None:
    """SAP 날짜 텍스트 'YYYY.MM.DD' → Python date. 빈값이면 None."""
    s = str(val).strip() if val is not None else ""
    if not s:
        return None
    try:
        parts = s.split(".")
        return date(int(parts[0]), int(parts[1]), int(parts[2]))
    except Exception:
        return None


def _parse_amount(val) -> float | None:
    """총계정원장 텍스트 → float. 빈값이면 None.
    지원 형식:
      '1,234.56'  — 영미식
      '1.234,56'  — 유럽식
      '1234.56-'  — 후행 마이너스 (일부 SAP 로케일)
    """
    s = str(val).strip() if val is not None else ""
    if not s:
        return None
    # 후행 마이너스 → 선행 마이너스로 정규화
    trailing_minus = s.endswith("-")
    if trailing_minus:
        s = "-" + s[:-1]
    try:
        cleaned = s.replace(",", "").replace(" ", "")
        return float(cleaned)
    except ValueError:
        pass
    # 유럽식: 마지막 ',' 가 소수점 구분자인 경우 ('1.234,56')
    try:
        cleaned = s.replace(".", "").replace(",", ".").replace(" ", "")
        return float(cleaned)
    except Exception:
        return None


def _adjust_formula(formula: str, row_offset: int) -> str:
    """수식의 상대 행 참조만 row_offset만큼 조정. 절대 참조($행)는 유지."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def replacer(m):
        col_ref   = m.group(1)   # 열 참조 (예: "A" 또는 "$A")
        dollar_row = m.group(2)  # 행 절대 참조 여부 ("$" 또는 "")
        row_num    = m.group(3)  # 행 번호 문자열
        if dollar_row:           # 절대 행 참조 → 변경 안 함
            return m.group(0)
        return f"{col_ref}{int(row_num) + row_offset}"

    return re.sub(r"(\$?[A-Z]+)(\$?)(\d+)", replacer, formula)


def _find_df_col(df: pd.DataFrame, keyword: str) -> str | None:
    """DataFrame 컬럼명 중 keyword로 시작하는 첫 번째 컬럼명 반환. 없으면 None.
    공백 정규화(strip spaces) 비교도 병행해 'col_net_due_date = 순만기일' ↔ '순 만기일' 불일치 대응."""
    norm_kw = keyword.replace(" ", "")
    for col in df.columns:
        col_s = str(col)
        if col_s.startswith(keyword):
            return col
        if col_s.replace(" ", "").startswith(norm_kw):
            return col
    return None


# 컬럼 역할별 헤더 이름 alias (한국어·영어·스페인어 혼용 파일 대응)
COL_ALIASES: dict[str, list[str]] = {
    "anchor":       ["지정", "DIV", "Sap code", "JE No."],
    "jijung":       ["지정", "DIV", "Sap code", "JE No."],
    "jeungbil":     ["증빙일", "DATE", "Date"],
    "elapsed":      ["경과기간", "Expiration period"],
    "text":         ["텍스트", "REMARK", "Text", "TEXT", "Contents"],
    "amount":       ["금액", "AMT", "Amount", "Amt.", "AMOUNT"],
    "currency":     ["통화", "CUR", "Currency"],
    "gigsanghwan":  ["기상환액", "상환액", "REPAYMENT", "Repayment", "Pay back", "REPAY"],
    "ar_balance":   ["상환 후 잔액", "SALDO", "Balance after repayment", "Balance", "BALANCE"],
    "sanghwanil":   ["상환일", "상환일(반제전표)", "PAYMENT DATE", "Repayment date"],
    "banjejeunpyo": ["반제전표", "반제전표번호", "Document number"],
    "mangil":       ["만기일", "EXPIRATION DATE"],
}


def _find_col_idx(ws, header_row: int, col_name: "str | list[str]") -> int | None:
    """헤더 행에서 col_name(단일 문자열 또는 alias 목록)과 일치하는 열 인덱스 반환.
    ws[header_row] 대신 ws.cell() 직접 접근으로 max_column 전체 탐색."""
    aliases = col_name if isinstance(col_name, list) else [col_name]
    max_col = max(ws.max_column or 0, 30)
    for c in range(1, max_col + 1):
        val = ws.cell(row=header_row, column=c).value
        if val and str(val).strip() in aliases:
            return c
    return None


def _find_header_row(ws, anchor_aliases: list[str], max_scan: int = 15) -> int | None:
    """anchor_aliases 중 하나가 있는 행 번호 반환. 없으면 None."""
    for r in range(1, max_scan + 1):
        for c in range(1, 30):
            val = ws.cell(r, c).value
            if val and str(val).strip() in anchor_aliases:
                return r
    return None


def _find_last_data_row(ws, header_row: int, anchor_col_idx: int) -> int:
    """anchor_col_idx 컬럼에서 마지막 데이터가 있는 행 번호 반환. 없으면 header_row."""
    for r in range(ws.max_row, header_row, -1):
        if ws.cell(row=r, column=anchor_col_idx).value is not None:
            return r
    return header_row


def _copy_row_format(ws, src_row: int, dst_row: int, max_col: int):
    """src_row 셀 서식(테두리·폰트·채우기·정렬·숫자형식)을 dst_row에 복사."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font        = _copy_obj(src.font)
            dst.border      = _copy_obj(src.border)
            dst.fill        = _copy_obj(src.fill)
            dst.alignment   = _copy_obj(src.alignment)
            dst.number_format = src.number_format


def _expand_formula_range(formula: str, new_end_row: int) -> str:
    """수식 내 범위 참조(A1:A100)의 끝 행이 new_end_row보다 작으면 확장."""
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    def replacer(m):
        end_row = int(m.group(4))
        if end_row >= new_end_row:
            return m.group(0)
        return f"{m.group(1)}:{m.group(2)}{m.group(3)}{new_end_row}"

    # 패턴: $?[A-Z]+$?\d+ : $?[A-Z]+ $? \d+
    return re.sub(
        r'(\$?[A-Z]+\$?\d+):(\$?[A-Z]+)(\$?)(\d+)',
        replacer,
        formula,
        flags=re.IGNORECASE,
    )


def _copy_cell_above(ws, new_row: int, col_idx: int, offset_from: int):
    """
    new_row 바로 위(new_row-1) 셀의 수식/값을 복사하여 new_row에 기입.
    수식이면 행 참조를 +1 조정, 값이면 그대로 복사.
    offset_from: 수식 행 참조 조정 기준이 되는 원본 행 (= new_row - 1)
    """
    src = ws.cell(row=new_row - 1, column=col_idx).value
    if src is None:
        return
    if isinstance(src, str) and src.startswith("="):
        ws.cell(row=new_row, column=col_idx).value = _adjust_formula(src, 1)
    else:
        ws.cell(row=new_row, column=col_idx).value = src


def append_to_source_file(df: pd.DataFrame, dest_file: Path, config, logger, yyyymm: str):
    """
    df를 SG 컬럼 기준으로 분류하여 dest_file 의 해당 시트에 지정 열만 append.

    열 매핑:
      지정   ← SAP col_jijung
      증빙일  ← SAP col_budat  (YYYY-MM-DD 날짜 변환, 반제일 있는 행 제외)
      경과기간← 기존 마지막 행 수식 복사
      텍스트 ← SAP col_text
      금액   ← SAP col_gl_amount (숫자 변환)
      통화   ← SAP col_currency

    시트별 헤더 행:
      외화외상매출금(잔액) → 8행,  나머지 → 4행
    외화외상매출금(잔액) D6 → 조회 월 말일로 설정
    """
    sg_col     = config.get("APPEND", "sg_column",          fallback="SG")
    gl_col     = config.get("APPEND", "gl_account_column",  fallback="G/L 계정")
    sh_mis_bal_candidates = [s.strip() for s in config.get("APPEND", "sheet_misugeun_bal", fallback="미수금(잔액)").split("|")]
    sh_mis_candidates     = [s.strip() for s in config.get("APPEND", "sheet_misugeun",     fallback="미수금").split("|")]
    sh_ar_bal_candidates  = [s.strip() for s in config.get("APPEND", "sheet_ar_bal",       fallback="외화외상매출금(잔액)").split("|")]
    sh_ar_candidates      = [s.strip() for s in config.get("APPEND", "sheet_ar",           fallback="외화외상매출금").split("|")]

    hdr_default = config.getint("APPEND", "header_row_default", fallback=4)
    hdr_ar_bal  = config.getint("APPEND", "header_row_ar_bal",  fallback=8)

    # SAP 컬럼명
    sap_jijung       = config.get("APPEND", "col_jijung",       fallback="지정")
    sap_budat        = config.get("APPEND", "col_budat",        fallback="전기일")
    sap_augdt        = config.get("APPEND", "col_augdt",        fallback="반제일")
    sap_text         = config.get("APPEND", "col_text",         fallback="텍스트")
    sap_gl_amount    = config.get("APPEND", "col_gl_amount",    fallback="총계정원장금액")
    sap_currency     = config.get("APPEND", "col_currency",     fallback="통화")
    sap_net_due_date = config.get("APPEND", "col_net_due_date", fallback="순만기일")

    formula_row_offsets = [int(s.strip()) for s in config.get("APPEND", "formula_rows_offset", fallback="1,2").split(",")]

    # SAP 컬럼명: 키워드 포함 방식으로 실제 컬럼명 탐색
    actual_sg_col     = _find_df_col(df, sg_col)
    actual_gl_col     = _find_df_col(df, gl_col)
    actual_augdt_col  = _find_df_col(df, sap_augdt)
    actual_jijung_col = _find_df_col(df, sap_jijung)
    actual_budat_col  = _find_df_col(df, sap_budat)
    actual_text_col   = _find_df_col(df, sap_text)
    actual_amount_col = _find_df_col(df, sap_gl_amount)
    actual_currency_col  = _find_df_col(df, sap_currency)
    actual_net_due_col   = _find_df_col(df, sap_net_due_date)

    logger.debug(f"  컬럼 매핑: SG={actual_sg_col}, 반제일={actual_augdt_col}, "
                 f"전기일={actual_budat_col}, 금액={actual_amount_col}")

    if actual_sg_col is None:
        logger.warning(f"  SG 컬럼 키워드 '{sg_col}' 없음 — append 건너뜀. 실제 컬럼: {list(df.columns)}")
        return

    # 반제일 있는 행 제외
    if actual_augdt_col:
        before = len(df)
        df = df[df[actual_augdt_col].astype(str).str.strip() == ""].reset_index(drop=True)
        skipped = before - len(df)
        if skipped:
            logger.info(f"  반제일 있는 행 {skipped}건 제외")

    # 분류
    mask_m  = df[actual_sg_col].astype(str).str.contains("M", na=False)
    mask_gl = (~mask_m & df[actual_gl_col].astype(str).str.strip().ne("")) \
              if actual_gl_col else ~mask_m

    df_mis = df[mask_m].reset_index(drop=True)
    df_ar  = df[mask_gl].reset_index(drop=True)
    logger.info(f"  분류 결과 — 미수금: {len(df_mis)}행 / 외화외상매출금: {len(df_ar)}행")

    wb = openpyxl.load_workbook(dest_file, keep_links=True)

    def resolve_sheet(candidates: list[str]) -> str | None:
        for name in candidates:
            if name in wb.sheetnames:
                return name
        return None

    for candidates, data, is_ar_bal, is_bal_sheet in [
        (sh_mis_bal_candidates, df_mis, False, True),
        (sh_mis_candidates,     df_mis, False, False),
        (sh_ar_bal_candidates,  df_ar,  True,  True),
        (sh_ar_candidates,      df_ar,  False, False),
    ]:
        sheet_name = resolve_sheet(candidates)
        if sheet_name is None:
            logger.warning(f"  시트 없음 (후보: {candidates}) — 건너뜀")
            continue
        if data.empty:
            logger.info(f"  [{sheet_name}] 추가할 데이터 없음")
            continue

        ws = wb[sheet_name]
        if is_ar_bal:
            header_row = _find_header_row(ws, COL_ALIASES["anchor"]) or hdr_ar_bal
        else:
            header_row = hdr_default

        # 헤더 행에서 대상 열 인덱스 수집 (alias 사전 사용)
        col_idx = {
            k: _find_col_idx(ws, header_row, COL_ALIASES[k])
            for k in ("jijung", "jeungbil", "elapsed", "text", "amount", "currency", "ar_balance", "gigsanghwan")
        }
        col_idx["mangil"] = _find_col_idx(ws, header_row, COL_ALIASES["mangil"]) if is_bal_sheet else None
        logger.debug(f"  [{sheet_name}] header_row={header_row}, 열 인덱스: {col_idx}")

        # 증빙일 컬럼 기준 마지막 데이터 행 찾기
        anchor_idx = col_idx["jeungbil"]
        if anchor_idx:
            last_data_row = _find_last_data_row(ws, header_row, anchor_idx)
        else:
            last_data_row = ws.max_row
        next_row = last_data_row + 1
        logger.debug(f"  [{sheet_name}] 마지막 데이터 행: {last_data_row}, 삽입 시작행: {next_row}")

        # 시트 최대 열 수 (서식 복사 범위용)
        max_col = ws.max_column

        # 행 기입 (위 셀 복사 대상: elapsed, ar_balance)
        for row_offset, (_, row) in enumerate(data.iterrows()):
            r = next_row + row_offset

            # 서식 복사: 마지막 데이터 행 서식을 새 행에 적용
            _copy_row_format(ws, last_data_row, r, max_col)

            if col_idx["jijung"] and actual_jijung_col:
                v = row.get(actual_jijung_col, "")
                ws.cell(row=r, column=col_idx["jijung"]).value = None if pd.isna(v) else v

            if col_idx["jeungbil"] and actual_budat_col:
                ws.cell(row=r, column=col_idx["jeungbil"]).value = _parse_sap_date(row.get(actual_budat_col))

            # elapsed는 append/offset 완료 후 apply_elapsed_formulas에서 일괄 처리

            if col_idx["text"] and actual_text_col:
                v = row.get(actual_text_col, "")
                ws.cell(row=r, column=col_idx["text"]).value = None if pd.isna(v) else v

            if col_idx["amount"] and actual_amount_col:
                ws.cell(row=r, column=col_idx["amount"]).value = _parse_amount(row.get(actual_amount_col))

            if col_idx["currency"] and actual_currency_col:
                v = row.get(actual_currency_col, "")
                ws.cell(row=r, column=col_idx["currency"]).value = None if pd.isna(v) else v

            # ar_balance 수식은 offset 삭제 완료 후 apply_elapsed_formulas에서 일괄 처리

            if col_idx["mangil"] and actual_net_due_col:
                _mc = ws.cell(row=r, column=col_idx["mangil"])
                if type(_mc).__name__ != "MergedCell":
                    _mc.value = _parse_sap_date(row.get(actual_net_due_col))  # type: ignore[assignment]

        added = len(data)
        final_data_row = next_row + added - 1

        # 수식 범위 확장: 금액·기상환액·상환 후 잔액 헤더 기준 1~2행 수식
        for role in ("amount", "gigsanghwan", "ar_balance"):
            fcol_idx = _find_col_idx(ws, header_row, COL_ALIASES[role])
            if not fcol_idx:
                continue
            for foffset in formula_row_offsets:
                frow = header_row + foffset
                cell = ws.cell(row=frow, column=fcol_idx)
                if isinstance(cell, openpyxl.cell.cell.MergedCell):
                    continue
                if cell.value and isinstance(cell.value, str) and cell.value.startswith("="):
                    expanded = _expand_formula_range(cell.value, final_data_row)
                    if expanded != cell.value:
                        cell.value = expanded
                        logger.debug(f"  [{sheet_name}] 수식 확장: {role} 행{frow} → {expanded}")

        logger.info(f"  [{sheet_name}] {added}행 추가 완료 (시작행: {next_row})")

    wb.save(dest_file)
    logger.info(f"  저장 완료: {dest_file.name}")


def append_offset_to_source_file(df_offset: pd.DataFrame, dest_file: Path, config, logger):
    """
    반제 데이터(offset df)를 채권명세서에 반영.

    2-1) 지정 열 매칭 → 기상환액에 총계정원장금액 기입 (모든 시트)
    2-2) 지정 열 매칭 → 상환일/반제전표 기입 (외화외상매출금/미수금 시트, 잔액 시트 제외)
    2-3) 잔액 시트 → 증빙일 있고 금액 = 기상환액(상환 후 잔액 = 0)인 행 삭제
    """
    sap_jijung = config.get("APPEND", "col_jijung",       fallback="지정")
    sap_amount = config.get("APPEND", "col_gl_amount",    fallback="총계정원장금액")
    sap_augdt  = config.get("APPEND", "col_augdt",        fallback="반제일")
    sap_augbl  = config.get("APPEND", "col_clearing_doc", fallback="반제전표")

    hdr_default = config.getint("APPEND", "header_row_default", fallback=4)
    hdr_ar_bal  = config.getint("APPEND", "header_row_ar_bal",  fallback=8)

    sh_mis_bal = [s.strip() for s in config.get("APPEND", "sheet_misugeun_bal", fallback="미수금(잔액)").split("|")]
    sh_mis     = [s.strip() for s in config.get("APPEND", "sheet_misugeun",     fallback="미수금").split("|")]
    sh_ar_bal  = [s.strip() for s in config.get("APPEND", "sheet_ar_bal",       fallback="외화외상매출금(잔액)").split("|")]
    sh_ar      = [s.strip() for s in config.get("APPEND", "sheet_ar",           fallback="외화외상매출금").split("|")]

    actual_jijung = _find_df_col(df_offset, sap_jijung)
    actual_amount = _find_df_col(df_offset, sap_amount)
    actual_augdt  = _find_df_col(df_offset, sap_augdt)
    actual_augbl  = _find_df_col(df_offset, sap_augbl)

    if actual_jijung is None:
        logger.warning(f"  [offset] 지정 컬럼 '{sap_jijung}' 없음 — offset append 건너뜀. 실제 컬럼: {list(df_offset.columns)}")
        return

    # 지정값 → 해당 offset 행 목록 조회용 딕셔너리
    # 키: 원본 문자열 AND 숫자 변환 값(앞 0 소멸) 동시 등록
    offset_lookup: dict[str, list] = {}
    for _, row in df_offset.iterrows():
        raw_key = str(row[actual_jijung]).strip()
        if not raw_key:
            continue
        offset_lookup.setdefault(raw_key, []).append(row)
        # 숫자로 해석 가능하면 int 변환 키도 등록 (앞 0 제거 대응)
        try:
            num_key = str(int(raw_key))
            if num_key != raw_key:
                offset_lookup.setdefault(num_key, []).append(row)
        except ValueError:
            pass

    if not offset_lookup:
        logger.info("  [offset] 매칭 가능한 지정값 없음")
        return

    actual_budat_offset = _find_df_col(df_offset, config.get("APPEND", "col_budat", fallback="전기일"))

    wb = openpyxl.load_workbook(dest_file, keep_links=True)

    def resolve_sheet(candidates: list[str]) -> str | None:
        for name in candidates:
            if name in wb.sheetnames:
                return name
        return None

    # (후보시트, ar_bal여부, 잔액시트여부, 상환일·반제전표 기입여부)
    for candidates, is_ar_bal_sheet, is_bal, write_extra in [
        (sh_mis_bal, False, True,  False),
        (sh_mis,     False, False, True),
        (sh_ar_bal,  True,  True,  False),
        (sh_ar,      False, False, True),
    ]:
        sheet_name = resolve_sheet(candidates)
        if sheet_name is None:
            logger.warning(f"  [offset] 시트 없음 (후보: {candidates}) — 건너뜀")
            continue

        ws = wb[sheet_name]
        if is_ar_bal_sheet:
            hdr = _find_header_row(ws, COL_ALIASES["anchor"]) or hdr_ar_bal
        else:
            hdr = hdr_default

        jijung_idx      = _find_col_idx(ws, hdr, COL_ALIASES["jijung"])
        jeungbil_idx    = _find_col_idx(ws, hdr, COL_ALIASES["jeungbil"])
        amount_idx      = _find_col_idx(ws, hdr, COL_ALIASES["amount"])
        gigsanghwan_idx = _find_col_idx(ws, hdr, COL_ALIASES["gigsanghwan"])
        sanghwanil_idx  = _find_col_idx(ws, hdr, COL_ALIASES["sanghwanil"])   if write_extra else None
        banjejeunpyo_idx= _find_col_idx(ws, hdr, COL_ALIASES["banjejeunpyo"]) if write_extra else None

        if not jijung_idx:
            logger.warning(f"  [offset] [{sheet_name}] 지정 열 없음 — 건너뜀")
            continue

        # 컬럼 탐색 결과 INFO 로그 (진단용)
        logger.info(
            f"  [offset] [{sheet_name}] header={hdr} | "
            f"지정={jijung_idx} 증빙일={jeungbil_idx} 금액={amount_idx} "
            f"기상환액={gigsanghwan_idx} 상환일={sanghwanil_idx} 반제전표={banjejeunpyo_idx}"
        )
        if not gigsanghwan_idx:
            # 헤더 행 실제 값 출력 (alias 불일치 진단용)
            actual_headers = [
                ws.cell(row=hdr, column=c).value
                for c in range(1, min((ws.max_column or 0) + 1, 40))
                if ws.cell(row=hdr, column=c).value is not None
            ]
            logger.warning(
                f"  [offset] [{sheet_name}] ⚠️ 기상환액 열 못 찾음 → 기상환액 기입 건너뜀. "
                f"실제 헤더: {actual_headers}"
            )

        anchor_idx    = jeungbil_idx or jijung_idx
        last_data_row = _find_last_data_row(ws, hdr, anchor_idx)

        # 2-1, 2-2: 지정 매칭 → 값 기입
        matched = 0
        for data_row in range(hdr + 1, last_data_row + 1):
            ws_jijung = ws.cell(row=data_row, column=jijung_idx).value
            if ws_jijung is None:
                continue

            # 지정값: 원본 문자열 AND 숫자 변환(앞 0 제거) 둘 다 시도
            raw_ws_key = str(ws_jijung).strip()
            try:
                num_ws_key = str(int(raw_ws_key))
            except ValueError:
                num_ws_key = raw_ws_key

            offset_rows = offset_lookup.get(raw_ws_key) or offset_lookup.get(num_ws_key)
            if not offset_rows:
                continue

            # 무조건 증빙일 일치 확인 (지정값 + 증빙일 둘 다 일치해야 반제로 인식)
            # openpyxl은 날짜 셀을 datetime으로 반환하므로 date로 정규화 후 비교
            if jeungbil_idx and actual_budat_offset:
                ws_jeungbil = ws.cell(row=data_row, column=jeungbil_idx).value
                if ws_jeungbil is not None:
                    if isinstance(ws_jeungbil, datetime):
                        ws_date = ws_jeungbil.date()
                    elif isinstance(ws_jeungbil, date):
                        ws_date = ws_jeungbil
                    else:
                        ws_date = _parse_sap_date(str(ws_jeungbil))
                    matched_rows = [
                        r for r in offset_rows
                        if _parse_sap_date(str(r.get(actual_budat_offset, ""))) == ws_date
                    ]
                    if not matched_rows:
                        continue  # 지정값 일치해도 증빙일 불일치 → 반제 아님
                    offset_rows = matched_rows

            # 2-1) 기상환액 = 매칭된 offset 행들의 총계정원장금액 합계
            if gigsanghwan_idx and actual_amount:
                total = sum((_parse_amount(r.get(actual_amount)) or 0.0) for r in offset_rows)
                ws.cell(row=data_row, column=gigsanghwan_idx).value = total  # type: ignore[assignment]

            # 2-2) 상환일 / 반제전표 (잔액 시트 제외)
            if write_extra:
                first = offset_rows[0]
                if sanghwanil_idx and actual_augdt:
                    ws.cell(row=data_row, column=sanghwanil_idx).value = _parse_sap_date(first.get(actual_augdt))  # type: ignore[assignment]
                if banjejeunpyo_idx and actual_augbl:
                    v = first.get(actual_augbl, "")
                    ws.cell(row=data_row, column=banjejeunpyo_idx).value = None if pd.isna(v) else v

            matched += 1

        logger.info(f"  [offset] [{sheet_name}] {matched}행 매칭 완료")

        # 2-3) 잔액 시트: 완전 수금 행 삭제 / 부분 수금 행 금액 차감
        if is_bal and amount_idx and gigsanghwan_idx and jeungbil_idx:
            last_row = _find_last_data_row(ws, hdr, anchor_idx)
            rows_to_delete = []
            partial_count = 0
            for data_row in range(hdr + 1, last_row + 1):
                if ws.cell(row=data_row, column=jeungbil_idx).value is None:
                    continue
                try:
                    amt = ws.cell(row=data_row, column=amount_idx).value
                    gig = ws.cell(row=data_row, column=gigsanghwan_idx).value
                    amt_f = float(amt) if amt not in (None, "") else 0.0  # type: ignore[arg-type]
                    gig_f = float(gig) if gig not in (None, "") else 0.0  # type: ignore[arg-type]
                    if abs(amt_f - gig_f) < 0.01:
                        # 완전 수금 → 행 삭제
                        rows_to_delete.append(data_row)
                    elif gig_f > 0:
                        # 부분 수금: 금액 -= 부분수금액, 기상환액 = 0 (상환후잔액 수식 정합성 유지)
                        partial_count += 1
                        logger.warning(
                            f"  [부분수금] [{sheet_name}] {data_row}행: "
                            f"원금={amt_f:,.2f}, 수금액={gig_f:,.2f}, 잔액={amt_f - gig_f:,.2f}"
                        )
                        ws.cell(row=data_row, column=amount_idx).value = amt_f - gig_f  # type: ignore[assignment]
                        ws.cell(row=data_row, column=gigsanghwan_idx).value = 0          # type: ignore[assignment]
                except (ValueError, TypeError):
                    pass
            for row_num in sorted(rows_to_delete, reverse=True):
                ws.delete_rows(row_num)
            if rows_to_delete:
                logger.info(f"  [offset] [{sheet_name}] {len(rows_to_delete)}행 삭제 (상환 후 잔액 = 0)")
            if partial_count:
                logger.warning(f"  ⚠️  [부분수금] [{sheet_name}] {partial_count}건 발생 — 금액 차감 처리 완료")

    wb.save(dest_file)
    logger.info(f"  [offset] 저장 완료: {dest_file.name}")


def apply_elapsed_formulas(dest_file: Path, config, budat_high_date: date, logger):
    """
    경과기간 수식을 append·offset 완료 후 마지막에 일괄 기입.
    - D6 셀에 budat_high_date 기입 (모든 시트)
    - 경과기간 비어 있는 새 행: =DATEDIF(증빙일셀, $D$6, "Y")
    """
    hdr_default = config.getint("APPEND", "header_row_default", fallback=4)
    hdr_ar_bal  = config.getint("APPEND", "header_row_ar_bal",  fallback=8)

    all_candidates = [
        ([s.strip() for s in config.get("APPEND", "sheet_misugeun_bal", fallback="미수금(잔액)").split("|")],  False),
        ([s.strip() for s in config.get("APPEND", "sheet_misugeun",     fallback="미수금").split("|")],        False),
        ([s.strip() for s in config.get("APPEND", "sheet_ar_bal",       fallback="외화외상매출금(잔액)").split("|")], True),
        ([s.strip() for s in config.get("APPEND", "sheet_ar",           fallback="외화외상매출금").split("|")], False),
    ]

    wb = openpyxl.load_workbook(dest_file, keep_links=True)

    modified = False
    for candidates, is_ar_bal in all_candidates:
        sheet_name = next((n for n in candidates if n in wb.sheetnames), None)
        if sheet_name is None:
            continue

        ws = wb[sheet_name]
        hdr = (_find_header_row(ws, COL_ALIASES["anchor"]) or hdr_ar_bal) if is_ar_bal else hdr_default

        elapsed_idx    = _find_col_idx(ws, hdr, COL_ALIASES["elapsed"])
        jeungbil_idx   = _find_col_idx(ws, hdr, COL_ALIASES["jeungbil"])
        jijung_idx     = _find_col_idx(ws, hdr, COL_ALIASES["jijung"])
        amount_idx     = _find_col_idx(ws, hdr, COL_ALIASES["amount"])
        gigsanghwan_idx= _find_col_idx(ws, hdr, COL_ALIASES["gigsanghwan"])
        ar_balance_idx = _find_col_idx(ws, hdr, COL_ALIASES["ar_balance"])
        anchor_idx     = jeungbil_idx or jijung_idx
        if not anchor_idx:
            continue

        last_data_row = _find_last_data_row(ws, hdr, anchor_idx)
        if last_data_row <= hdr:
            continue

        # anchor 값 있는 모든 데이터 행 (삭제 후 실제 행 번호 기준)
        all_data_rows = [
            r for r in range(hdr + 1, last_data_row + 1)
            if ws.cell(row=r, column=anchor_idx).value is not None
        ]
        if not all_data_rows:
            continue

        # D6에 조회 종료일 기입
        _d6 = ws.cell(row=6, column=4)
        if type(_d6).__name__ != "MergedCell":
            _d6.value = budat_high_date  # type: ignore[assignment]
            logger.info(f"  [{sheet_name}] D6 → {budat_high_date} 기입")

        # 경과기간 수식 전체 재기입: =DATEDIF(증빙일셀, $D$6, "Y")
        if elapsed_idx and jeungbil_idx:
            jeungbil_col = get_column_letter(jeungbil_idx)
            cnt = 0
            for r in all_data_rows:
                _cell = ws.cell(row=r, column=elapsed_idx)
                if type(_cell).__name__ != "MergedCell":
                    _cell.value = f'=DATEDIF({jeungbil_col}{r},$D$6,"Y")'  # type: ignore[assignment]
                    cnt += 1
            logger.info(f"  [{sheet_name}] 경과기간 수식 {cnt}행 기입")
            modified = True

        # 상환 후 잔액 수식 전체 재기입: =금액셀 - 기상환액셀
        if ar_balance_idx and amount_idx and gigsanghwan_idx:
            amt_col = get_column_letter(amount_idx)
            gig_col = get_column_letter(gigsanghwan_idx)
            cnt = 0
            for r in all_data_rows:
                _cell = ws.cell(row=r, column=ar_balance_idx)
                if type(_cell).__name__ != "MergedCell":
                    _cell.value = f"={amt_col}{r}-{gig_col}{r}"  # type: ignore[assignment]
                    cnt += 1
            logger.info(f"  [{sheet_name}] 상환 후 잔액 수식 {cnt}행 기입")
            modified = True

    if modified:
        wb.save(dest_file)


def get_customer_accounts(source_dir: Path, logger) -> list[str]:
    """source_dir 의 파일명이 7자리 숫자로 시작하는 파일에서 고객계정 수집"""
    import re
    accounts = []
    seen = set()
    for f in sorted(source_dir.iterdir()):
        if not f.is_file():
            continue
        stem = f.stem
        # 파일명이 7자리 숫자로 시작하는 경우만 추출
        match = re.match(r'^(\d{7})', stem)
        if not match:
            logger.debug(f"고객계정 아님 — 건너뜀: {f.name}")
            continue
        account = match.group(1)
        if account not in seen:
            seen.add(account)
            accounts.append(account)
            logger.info(f"고객계정 추출: {account}  ← {f.name}")
    return accounts


# ──────────────────────────────────────────────────────────
# FBL5N 실행 클래스
# ──────────────────────────────────────────────────────────

class FBL5NDownloader:
    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.session: "Any" = None

        env_path = Path(__file__).parent / "config" / ".env"
        load_dotenv(dotenv_path=env_path)

        # config.ini 에서 필드 ID 로드
        self.transaction      = config.get("SAP", "transaction", fallback="FBL5N")
        self.customer_field     = config.get("SAP", "customer_field_id")
        self.company_code       = config.get("SAP", "company_code", fallback="1000")
        self.company_code_field = config.get("SAP", "company_code_field")
        self.all_items_radio    = config.get("SAP", "all_items_radio")
        self.budat_low_field    = config.get("SAP", "budat_low_field")
        self.budat_high_field   = config.get("SAP", "budat_high_field")
        self.normal_items_chk   = config.get("SAP", "normal_items_chk")
        self.special_gl_chk     = config.get("SAP", "special_gl_chk")
        self.noted_items_chk    = config.get("SAP", "noted_items_chk")
        self.posting_date_col   = config.get("SAP", "posting_date_col", fallback="BUDAT")
        self.execute_vkey       = config.getint("SAP", "execute_vkey", fallback=8)

        self.augdt_low_field      = config.get("SAP", "augdt_low_field",      fallback="wnd[0]/usr/ctxtSO_AUGDT-LOW")
        self.augdt_high_field     = config.get("SAP", "augdt_high_field",     fallback="wnd[0]/usr/ctxtSO_AUGDT-HIGH")
        self.cleared_items_radio  = config.get("SAP", "cleared_items_radio",  fallback="wnd[0]/usr/radX_CLSEL")

        self.raw_dir    = Path(config.get("PATHS", "raw_dir"))
        self.source_dir = Path(config.get("PATHS", "source_dir"))
        self.raw_dir.mkdir(parents=True, exist_ok=True)

    def connect(self):
        """실행 중인 SAP GUI 세션에 연결"""
        sap_gui_auto = win32com.client.GetObject("SAPGUI")
        application  = sap_gui_auto.GetScriptingEngine
        connection   = application.Children(0)
        self.session = connection.Children(0)
        self.logger.info("SAP GUI 세션 연결 완료")

    def run_all(self, accounts: list[str], budat_low: str, budat_high: str, yyyymm: str):
        """모든 고객계정에 대해 FBL5N 실행 → 미결항목 다운로드 → 반제항목 다운로드"""
        success, failed = [], []

        for i, account in enumerate(accounts, 1):
            self.logger.info(f"[{i}/{len(accounts)}] 계정: {account}  전기일: {budat_low} ~ {budat_high}")

            # 원본 파일 확인 + 복사본 생성 (이후 모든 작업은 복사본에만)
            source_file = find_source_file(self.source_dir, account)
            if source_file:
                working_copy = make_working_copy(source_file, account, budat_high)
                self.logger.info(f"  복사본 생성: {working_copy.name}")
            else:
                working_copy = None
                self.logger.warning(f"  원본 파일 없음 — append 건너뜀 (계정: {account})")

            # 1. 미결항목 (전기일 기간)
            try:
                dest = self._run_single(account, budat_low, budat_high, yyyymm, working_copy)
                self.logger.info(f"  → 저장 완료: {dest.name}")
            except Exception as e:
                self.logger.error(f"  → 실패 ({account}): {e}")
                failed.append(account)
                self._go_back_to_start()
                continue

            # 2. 반제항목 (반제일 기간) — 실패해도 계정 전체 실패 처리 안 함
            try:
                self._run_single_offset(account, budat_low, budat_high, yyyymm, working_copy)
            except Exception as e:
                self.logger.warning(f"  [offset] 건너뜀 ({account}): {e}")
                self._go_back_to_start()

            success.append(account)

        self.logger.info("=" * 50)
        self.logger.info(f"완료: 성공 {len(success)}건 / 실패 {len(failed)}건")
        if failed:
            self.logger.warning(f"실패 계정: {failed}")

    def _run_single(self, account: str, budat_low: str, budat_high: str, yyyymm: str,
                    working_copy: "Path | None") -> Path:
        """단일 고객계정 FBL5N 실행 → 저장 → 파일 경로 반환"""
        self._navigate_to_fbl5n()
        self._fill_selection_screen(account, budat_low, budat_high)
        self.session.findById("wnd[0]").sendVKey(self.execute_vkey)
        time.sleep(3)
        self.logger.info("  FBL5N 조회 실행")

        dest = self.raw_dir / f"{account}-{yyyymm}.xlsx"
        df = self._read_grid_and_save(dest)

        if working_copy:
            append_to_source_file(df, working_copy, self.config, self.logger, yyyymm)
        return dest

    def _run_single_offset(self, account: str, budat_low: str, budat_high: str, yyyymm: str,
                           working_copy: "Path | None") -> Path:
        """반제일 기간 기준 FBL5N 조회 → _offset.xlsx 저장 → offset append"""
        self._navigate_to_fbl5n()
        self._fill_selection_screen_offset(account, budat_low, budat_high)
        self.session.findById("wnd[0]").sendVKey(self.execute_vkey)
        time.sleep(3)
        self.logger.info("  [offset] FBL5N 조회 실행")

        dest = self.raw_dir / f"{account}-{yyyymm}_offset.xlsx"
        df_offset = self._read_grid_and_save(dest)

        if working_copy:
            append_offset_to_source_file(df_offset, working_copy, self.config, self.logger)
            # 경과기간 수식: append + offset 완료 후 마지막에 일괄 기입
            budat_high_date = _parse_sap_date(budat_high) or date.today()
            apply_elapsed_formulas(working_copy, self.config, budat_high_date, self.logger)
        else:
            self.logger.warning(f"  [offset] 복사본 없음 — offset append 건너뜀 (계정: {account})")

        return dest

    def _fill_selection_screen_offset(self, account: str, budat_low: str, budat_high: str):
        """FBL5N offset 선택 화면: 반제일 기간 입력, 전기일 비움"""
        s = self.session
        s.findById(self.customer_field).text = account
        try:
            s.findById(self.company_code_field).text = self.company_code
        except Exception as e:
            self.logger.warning(f"  [offset] 회사코드 입력 실패: {e}")
        # 반제항목 라디오 선택 (반제일 필터가 올바르게 동작하려면 반드시 반제항목 모드여야 함)
        try:
            s.findById(self.cleared_items_radio).select()
        except Exception as e:
            self.logger.warning(f"  [offset] 반제항목 라디오 실패: {e}")
        # 전기일 초기화 (SAP이 이전 값 기억할 수 있음)
        try:
            s.findById(self.budat_low_field).text  = ""
            s.findById(self.budat_high_field).text = ""
        except Exception:
            pass
        # 반제일 기간 입력
        try:
            s.findById(self.augdt_low_field).text  = budat_low
            s.findById(self.augdt_high_field).text = budat_high
        except Exception as e:
            self.logger.warning(f"  [offset] 반제일 기간 입력 실패: {e}")
        for chk, name in [
            (self.normal_items_chk, "일반항목"),
            (self.special_gl_chk,   "특별G/L"),
            (self.noted_items_chk,  "임시항목"),
        ]:
            try:
                s.findById(chk).selected = True
            except Exception as e:
                self.logger.warning(f"  [offset] {name} 체크 실패: {e}")
        self.logger.info(f"  [offset] 선택 화면 완료 (계정: {account}, 반제일: {budat_low} ~ {budat_high})")

    def _navigate_to_fbl5n(self):
        """FBL5N 선택 화면으로 이동"""
        cmd = self.session.findById("wnd[0]/tbar[0]/okcd")
        cmd.text = f"/n{self.transaction}"
        self.session.findById("wnd[0]").sendVKey(0)
        time.sleep(2)

    def _fill_selection_screen(self, account: str, budat_low: str, budat_high: str):
        """FBL5N 선택 화면: 고객계정, 전기일 기간, 체크박스 입력"""
        s = self.session

        # 고객계정
        s.findById(self.customer_field).text = account

        # 회사코드
        try:
            s.findById(self.company_code_field).text = self.company_code
        except Exception as e:
            self.logger.warning(f"  회사코드 입력 실패: {e}")

        # 모든 항목 라디오 선택
        try:
            s.findById(self.all_items_radio).select()
        except Exception as e:
            self.logger.warning(f"  모든 항목 라디오 선택 실패: {e}")

        # 전기일 기간
        s.findById(self.budat_low_field).text  = budat_low
        s.findById(self.budat_high_field).text = budat_high

        # 일반항목 체크
        try:
            s.findById(self.normal_items_chk).selected = True
        except Exception as e:
            self.logger.warning(f"  일반항목 체크 실패: {e}")

        # 특별G/L거래 체크
        try:
            s.findById(self.special_gl_chk).selected = True
        except Exception as e:
            self.logger.warning(f"  특별G/L 체크 실패: {e}")

        # 임시항목 체크
        try:
            s.findById(self.noted_items_chk).selected = True
        except Exception as e:
            self.logger.warning(f"  임시항목 체크 실패: {e}")

        self.logger.info(f"  선택 화면 입력 완료 (계정: {account}, 전기일: {budat_low} ~ {budat_high})")

    def _read_grid_and_save(self, dest: Path) -> pd.DataFrame:
        """ALV 그리드 직접 읽기 → 전기일자 오름차순 정렬 → 엑셀 저장 → DataFrame 반환"""

        grid_id = self.config.get("SAP", "grid_id")
        grid = self.session.findById(grid_id)

        row_count = grid.RowCount
        self.logger.info(f"  그리드 행 수: {row_count}")

        if row_count == 0:
            raise ValueError("조회 결과가 없습니다.")

        # 컬럼 목록
        columns = list(grid.ColumnOrder)
        self.logger.info(f"  컬럼 수: {len(columns)}")

        # 컬럼 헤더(화면 표시명) 수집
        headers = {}
        for col in columns:
            try:
                headers[col] = grid.GetDisplayedColumnTitle(col)
            except Exception:
                headers[col] = col

        # 전체 데이터 읽기
        records = []
        for row in range(row_count):
            record = {}
            for col in columns:
                try:
                    record[col] = grid.GetCellValue(row, col)
                except Exception:
                    record[col] = ""
            records.append(record)

        df = pd.DataFrame(records, columns=columns)

        # 컬럼명을 화면 표시명으로 변경
        df = df.rename(columns=headers)

        # 전기일자 오름차순 정렬 (BUDAT 또는 표시명)
        budat_col = headers.get(self.posting_date_col, self.posting_date_col)
        if budat_col in df.columns:
            df = df.sort_values(budat_col).reset_index(drop=True)
            self.logger.info(f"  전기일자 오름차순 정렬 완료")
        else:
            self.logger.warning(f"  전기일자 컬럼 '{budat_col}' 없음 — 정렬 생략")

        # 엑셀 저장
        dest.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(dest, index=False)
        self.logger.info(f"  엑셀 저장 완료: {dest}  ({len(df)}행)")

        return df

    def _go_back_to_start(self):
        """오류 발생 시 초기 화면으로 복귀"""
        try:
            self.session.findById("wnd[0]/tbar[0]/okcd").text = "/n"
            self.session.findById("wnd[0]").sendVKey(0)
            time.sleep(1)
        except Exception:
            pass

    def close(self):
        self._go_back_to_start()


# ──────────────────────────────────────────────────────────
# 메인
# ──────────────────────────────────────────────────────────

def parse_args():
    parser = argparse.ArgumentParser(description="FBL5N 채권 미결항목 다운로드")
    parser.add_argument(
        "--keydate",
        required=False,
        help="조회 기준 년월 (예: 202503). 해당 월 말일로 자동 변환됩니다. --budat_low/high 와 동시 사용 불가.",
    )
    parser.add_argument(
        "--budat_low",
        help="전기일 시작 (YYYYMMDD). 예: 20260101",
    )
    parser.add_argument(
        "--budat_high",
        help="전기일 종료 (YYYYMMDD). 예: 20260331",
    )
    parser.add_argument(
        "--accounts",
        nargs="+",
        help="처리할 고객계정 목록 (미지정 시 source_dir 전체). 예: --accounts 1700006 1700051",
    )
    parser.add_argument(
        "--source_dir",
        help="채권명세서 파일 경로 (지정 시 config.ini source_dir 덮어씀, raw_dir은 하위 raw/ 로 자동 설정)",
    )
    return parser.parse_args()


def _parse_date_arg(s: str) -> str:
    """'20260101' → '2026.01.01' (SAP 날짜 형식)"""
    s = s.strip()
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}.{s[4:6]}.{s[6:]}"
    raise ValueError(f"날짜 형식 오류: {s!r}  (YYYYMMDD 형식으로 입력)")


def main():
    args = parse_args()

    if args.keydate and (args.budat_low or args.budat_high):
        print("ERROR: --keydate 와 --budat_low/--budat_high 는 동시에 사용할 수 없습니다.")
        sys.exit(1)
    if not args.keydate and not (args.budat_low and args.budat_high):
        print("ERROR: --keydate 또는 --budat_low + --budat_high 중 하나를 지정하세요.")
        sys.exit(1)

    if args.keydate:
        yyyymm     = args.keydate
        budat_low  = month_start(yyyymm)
        budat_high = month_end(yyyymm)
    else:
        budat_low  = _parse_date_arg(args.budat_low)
        budat_high = _parse_date_arg(args.budat_high)
        # yyyymm: raw 파일명 접미사로 사용 — 종료월 기준
        yyyymm = args.budat_high.strip()[:6]

    config = get_config()
    logger = setup_logger("sapost.fbl5n", config)

    # --source_dir 지정 시 config 메모리 내 덮어쓰기 (config.ini 파일 수정 없음)
    if args.source_dir:
        src = Path(args.source_dir)
        config.set("PATHS", "source_dir", str(src))
        config.set("PATHS", "raw_dir",    str(src / "raw"))
        logger.info(f"source_dir override: {src}")

    logger.info("=" * 60)
    logger.info(f"FBL5N 다운로드 시작 | 전기일: {budat_low} ~ {budat_high}")
    logger.info("=" * 60)

    # 고객계정 목록 수집
    source_dir = Path(config.get("PATHS", "source_dir"))
    if not source_dir.exists():
        logger.error(f"source_dir 를 찾을 수 없습니다: {source_dir}")
        sys.exit(1)

    if args.accounts:
        accounts = args.accounts
        logger.info(f"지정 계정 {len(accounts)}개: {accounts}")
    else:
        accounts = get_customer_accounts(source_dir, logger)
        if not accounts:
            logger.error("고객계정을 추출할 파일이 없습니다.")
            sys.exit(1)
        logger.info(f"총 {len(accounts)}개 계정 추출: {accounts}")

    # SAP 연결 및 실행
    downloader = FBL5NDownloader(config, logger)
    try:
        downloader.connect()
        downloader.run_all(accounts, budat_low, budat_high, yyyymm)
    finally:
        downloader.close()


if __name__ == "__main__":
    main()
