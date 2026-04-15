"""
Excel 비교 리포트 작성: 요약 시트, 변경내역 시트, 비교_[시트명] 상세 시트
"""
from __future__ import annotations

import configparser
import logging
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)
from openpyxl.utils import get_column_letter

from .comparator import WorkbookDiff, SheetDiff, CellChange, RowChange
from .excel_reader import SheetData


# ---------------------------------------------------------------------------
# 색상 상수 (기존 모듈과 통일)
# ---------------------------------------------------------------------------
COLOR_HEADER        = "1F4E79"   # 진파랑 — 헤더 배경
COLOR_HEADER_FONT   = "FFFFFF"   # 흰색 — 헤더 글꼴
COLOR_SUMMARY_FILL  = "BDD7EE"   # 연파랑 — 요약 행
COLOR_CHANGED_CELL  = "FFFF00"   # 노랑 — 변경된 셀
COLOR_ADDED_ROW     = "C6EFCE"   # 연초록 — 추가된 행
COLOR_DELETED_ROW   = "FFC7CE"   # 연빨강 — 삭제된 행
COLOR_ADDED_COL     = "E2EFDA"   # 연녹색 — 추가된 열
COLOR_REMOVED_COL   = "FCE4D6"   # 연주황 — 삭제된 열
COLOR_TOTAL_ROW     = "D9D9D9"   # 회색 — 합계 행


class ReportWriter:
    def __init__(self, config: configparser.ConfigParser, logger: logging.Logger) -> None:
        self.config = config
        self.logger = logger
        self.summary_name  = config.get("REPORT", "summary_sheet_name",  fallback="요약")
        self.changes_name  = config.get("REPORT", "changes_sheet_name",  fallback="변경내역")
        self.detail_prefix = config.get("REPORT", "detail_sheet_prefix", fallback="비교_")
        # 변경내역 컬럼 레이블
        self.col_sheet   = config.get("REPORT", "col_sheet",   fallback="Sheet")
        self.col_rowkey  = config.get("REPORT", "col_rowkey",  fallback="행키")
        self.col_colname = config.get("REPORT", "col_colname", fallback="열명")
        self.col_old     = config.get("REPORT", "col_old",     fallback="구버전")
        self.col_new     = config.get("REPORT", "col_new",     fallback="신버전")
        self.col_delta   = config.get("REPORT", "col_delta",   fallback="증감")
        self.col_pct     = config.get("REPORT", "col_pct",     fallback="증감률")

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def write(self,
              diff: WorkbookDiff,
              old_sheets: dict[str, SheetData],
              new_sheets: dict[str, SheetData],
              output_path: Path) -> None:
        wb = Workbook()
        # 기본 시트 제거
        wb.remove(wb.active)

        self._write_summary_sheet(wb, diff)
        self._write_changes_sheet(wb, diff)

        for sd in diff.sheet_diffs:
            if sd.old_only_sheet or sd.new_only_sheet or sd.skipped:
                continue
            old_sd = old_sheets.get(sd.sheet_name)
            new_sd = new_sheets.get(sd.sheet_name)
            self._write_detail_sheet(wb, sd, old_sd, new_sd)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        self.logger.info(f"리포트 저장 완료: {output_path}")

    # ------------------------------------------------------------------
    # 요약 시트
    # ------------------------------------------------------------------

    def _write_summary_sheet(self, wb: Workbook, diff: WorkbookDiff) -> None:
        ws = wb.create_sheet(self.summary_name)
        headers = ["시트명", "변경셀", "추가행", "삭제행", "추가열", "삭제열", "비고"]
        self._write_header_row(ws, 1, headers)

        total = [0, 0, 0, 0, 0]
        for r, row in enumerate(diff.summary_rows, start=2):
            vals = [
                row["sheet"],
                row["cell_changes"],
                row["added_rows"],
                row["deleted_rows"],
                row["added_cols"],
                row["removed_cols"],
                row["note"],
            ]
            for c, v in enumerate(vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = Alignment(horizontal="center" if c > 1 else "left",
                                           vertical="center")
                cell.border = self._thin_border()
                if row["note"] in ("(삭제됨)", "(추가됨)"):
                    cell.fill = self._fill(COLOR_SUMMARY_FILL)
            for i, key in enumerate(["cell_changes", "added_rows", "deleted_rows",
                                      "added_cols", "removed_cols"]):
                total[i] += row[key]

        # 합계 행
        tot_row = ws.max_row + 1
        ws.cell(row=tot_row, column=1, value="합계").font = Font(bold=True)
        for i, v in enumerate(total, start=2):
            cell = ws.cell(row=tot_row, column=i, value=v)
            cell.font = Font(bold=True)
            cell.fill = self._fill(COLOR_TOTAL_ROW)
            cell.border = self._thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=tot_row, column=1).fill = self._fill(COLOR_TOTAL_ROW)
        ws.cell(row=tot_row, column=1).border = self._thin_border()

        self._set_column_widths(ws, [25, 10, 10, 10, 10, 10, 12])
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # 변경내역 시트
    # ------------------------------------------------------------------

    def _write_changes_sheet(self, wb: Workbook, diff: WorkbookDiff) -> None:
        ws = wb.create_sheet(self.changes_name)
        headers = [self.col_sheet, self.col_rowkey, self.col_colname,
                   self.col_old, self.col_new, self.col_delta, self.col_pct]
        self._write_header_row(ws, 1, headers)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

        for r, ch in enumerate(diff.all_cell_changes, start=2):
            row_vals = [
                ch.sheet,
                ch.row_key,
                ch.col_name,
                ch.old_val,
                ch.new_val,
                ch.delta,
                ch.delta_pct,
            ]
            for c, v in enumerate(row_vals, start=1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = self._thin_border()
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=(c == 2))
            # 증감률 포맷
            if ch.delta_pct is not None:
                pct_cell = ws.cell(row=r, column=7)
                pct_cell.number_format = "0.00%"
                pct_cell.value = ch.delta_pct / 100  # openpyxl은 0.05 = 5%

        self._set_column_widths(ws, [18, 30, 20, 15, 15, 12, 10])
        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # 상세 비교 시트 (비교_[시트명])
    # ------------------------------------------------------------------

    def _write_detail_sheet(self,
                             wb: Workbook,
                             sd: SheetDiff,
                             old_sd: Optional[SheetData],
                             new_sd: Optional[SheetData]) -> None:
        sheet_title = f"{self.detail_prefix}{sd.sheet_name}"
        # 시트명 31자 제한
        if len(sheet_title) > 31:
            sheet_title = sheet_title[:31]
        ws = wb.create_sheet(sheet_title)

        if new_sd is None or new_sd.df.empty:
            ws.cell(row=1, column=1, value="데이터 없음")
            return

        df = new_sd.df

        # 표시할 컬럼 순서: 공통 → 추가된 열 → 삭제된 열(표시용)
        display_cols = list(df.columns)
        removed_col_set = set(sd.removed_cols)
        added_col_set = set(sd.added_cols)

        # 삭제된 열도 오른쪽에 추가 (구버전에서 가져와 표시)
        extra_deleted_cols: list[str] = []
        if old_sd is not None:
            for rc in sd.removed_cols:
                if rc in old_sd.df.columns and rc not in display_cols:
                    extra_deleted_cols.append(rc)
        all_display_cols = display_cols + extra_deleted_cols

        # 헤더 작성
        for c, col_name in enumerate(all_display_cols, start=1):
            cell = ws.cell(row=1, column=c, value=col_name)
            cell.font = Font(bold=True, color=COLOR_HEADER_FONT)
            cell.border = self._thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_name in removed_col_set:
                cell.fill = self._fill(COLOR_REMOVED_COL)
                cell.font = Font(bold=True, color="000000", strike=True)
            elif col_name in added_col_set:
                cell.fill = self._fill(COLOR_ADDED_COL)
            else:
                cell.fill = self._fill(COLOR_HEADER)

        # 변경 셀/행 빠른 조회용 셋
        changed_cell_set: set[tuple[str, str]] = {
            (ch.row_key, ch.col_name) for ch in sd.cell_changes
        }
        added_key_set = {r.row_key for r in sd.added_rows}
        deleted_key_set = {r.row_key for r in sd.deleted_rows}

        def _row_key(row_series) -> str:
            return " | ".join(str(row_series.get(k, "")) for k in new_sd.key_columns)

        # 신버전 데이터 행 작성
        current_row = 2
        for _, row in df.iterrows():
            rk = _row_key(row)
            is_added = rk in added_key_set
            row_fill = self._fill(COLOR_ADDED_ROW) if is_added else None

            for c, col_name in enumerate(all_display_cols, start=1):
                if col_name in extra_deleted_cols:
                    # 삭제된 열 — 구버전 값 표시
                    val = ""
                    if old_sd is not None and not old_sd.df.empty:
                        old_key_map = {
                            " | ".join(str(old_sd.df.iloc[i].get(k, ""))
                                       for k in old_sd.key_columns): i
                            for i in range(len(old_sd.df))
                        }
                        if rk in old_key_map:
                            val = old_sd.df.iloc[old_key_map[rk]].get(col_name, "")
                    cell = ws.cell(row=current_row, column=c, value=val)
                    cell.fill = self._fill(COLOR_REMOVED_COL)
                else:
                    val = row.get(col_name, "")
                    cell = ws.cell(row=current_row, column=c, value=val)
                    if row_fill:
                        cell.fill = row_fill
                    elif (rk, col_name) in changed_cell_set:
                        cell.fill = self._fill(COLOR_CHANGED_CELL)
                        # 코멘트: 구버전 값 표시
                        old_val = next(
                            (ch.old_val for ch in sd.cell_changes
                             if ch.row_key == rk and ch.col_name == col_name), ""
                        )
                        if old_val:
                            from openpyxl.comments import Comment
                            cell.comment = Comment(f"구버전: {old_val}", "VerCheck")

                cell.border = self._thin_border()
                cell.alignment = Alignment(vertical="center")

            current_row += 1

        # 삭제된 행 추가 (빨간 배경)
        if old_sd is not None:
            del_row_map = {r.row_key: r for r in sd.deleted_rows}
            for _, old_row in old_sd.df.iterrows():
                rk = " | ".join(str(old_row.get(k, "")) for k in old_sd.key_columns)
                if rk not in del_row_map:
                    continue
                for c, col_name in enumerate(all_display_cols, start=1):
                    if col_name in extra_deleted_cols:
                        val = old_row.get(col_name, "")
                    else:
                        val = old_row.get(col_name, "")
                    cell = ws.cell(row=current_row, column=c, value=val)
                    cell.fill = self._fill(COLOR_DELETED_ROW)
                    cell.border = self._thin_border()
                    cell.alignment = Alignment(vertical="center")
                current_row += 1

        # 범례
        legend_col = len(all_display_cols) + 2
        legends = [
            (COLOR_CHANGED_CELL, "값 변경"),
            (COLOR_ADDED_ROW,    "추가된 행"),
            (COLOR_DELETED_ROW,  "삭제된 행"),
            (COLOR_ADDED_COL,    "추가된 열"),
            (COLOR_REMOVED_COL,  "삭제된 열"),
        ]
        ws.cell(row=1, column=legend_col, value="범례").font = Font(bold=True)
        for i, (color, label) in enumerate(legends, start=2):
            cell = ws.cell(row=i, column=legend_col, value=label)
            cell.fill = self._fill(color)
            cell.border = self._thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 열 너비 균일 설정
        for c in range(1, len(all_display_cols) + 3):
            ws.column_dimensions[get_column_letter(c)].width = 16

        ws.freeze_panes = "A2"

    # ------------------------------------------------------------------
    # 스타일 헬퍼
    # ------------------------------------------------------------------

    def _write_header_row(self, ws, row: int, headers: list[str]) -> None:
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = Font(bold=True, color=COLOR_HEADER_FONT)
            cell.fill = self._fill(COLOR_HEADER)
            cell.border = self._thin_border()
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _fill(self, hex_color: str) -> PatternFill:
        return PatternFill(fill_type="solid", fgColor=hex_color)

    def _thin_border(self) -> Border:
        thin = Side(style="thin")
        return Border(left=thin, right=thin, top=thin, bottom=thin)

    def _set_column_widths(self, ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
