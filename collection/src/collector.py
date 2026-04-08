"""
영업수금(Collection) 집계 모듈

자금일보 엑셀 파일의 각 날짜 시트에서 inflow 테이블을 탐색하여
Category = 'Collection' 인 행의 Amount 합계(영업수금)를 산출한다.
"""
from __future__ import annotations

import re
import logging
from pathlib import Path
from dataclasses import dataclass, field

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


@dataclass
class CollectionRow:
    sheet: str
    row_num: int
    account_code: str | None
    customer: str | None
    category: str
    details_bank: str | None
    details_company: str | None
    amount: float


@dataclass
class CollectionResult:
    rows: list[CollectionRow] = field(default_factory=list)
    total: float = 0.0

    def by_sheet(self) -> dict[str, float]:
        result: dict[str, float] = {}
        for r in self.rows:
            result[r.sheet] = result.get(r.sheet, 0.0) + r.amount
        return result


class CollectionCollector:
    def __init__(self, config, logger_: logging.Logger | None = None):
        self.config = config
        self.logger = logger_ or logger

        self.target_category = config.get(
            "COLLECTION", "target_category", fallback="Collection"
        ).strip().lower()
        self.sheet_pattern = re.compile(
            config.get("COLLECTION", "sheet_pattern", fallback=r"^\d{2}\.\d{2}\.\d{4}")
        )
        self.header_scan_rows = config.getint(
            "COLLECTION", "header_scan_rows", fallback=60
        )

    # ──────────────────────────────────────────
    # 핵심 집계
    # ──────────────────────────────────────────

    def collect(self, source_file: Path) -> CollectionResult:
        self.logger.info(f"파일 로드: {source_file.name}")
        wb = openpyxl.load_workbook(source_file, keep_links=True, data_only=True)

        date_sheets = [s for s in wb.sheetnames if self.sheet_pattern.match(s)]
        self.logger.info(f"날짜 시트 수: {len(date_sheets)}")

        result = CollectionResult()

        for sheet_name in date_sheets:
            ws = wb[sheet_name]
            rows = self._collect_sheet(ws, sheet_name)
            if rows:
                result.rows.extend(rows)
                sheet_total = sum(r.amount for r in rows)
                self.logger.info(f"  [{sheet_name}] {len(rows)}건  합계: {sheet_total:,.2f}")
            else:
                self.logger.debug(f"  [{sheet_name}] Collection 없음")

        result.total = sum(r.amount for r in result.rows)
        self.logger.info(f"영업수금 합계: {result.total:,.2f}  ({len(result.rows)}건)")
        return result

    def _collect_sheet(self, ws, sheet_name: str) -> list[CollectionRow]:
        """한 시트에서 Category=Collection 행 수집"""
        hdr_row, cat_col = self._find_header(ws)
        if hdr_row is None:
            self.logger.warning(f"  [{sheet_name}] 헤더 행 없음 — 건너뜀")
            return []

        # inflow 테이블 열 위치: cat_col 기준
        # 헤더: ACCOUNT CODE | CUSTOMER | Category | Details(Bank) | Details(Company) | AMOUNT
        amt_col = cat_col + 3

        rows: list[CollectionRow] = []
        for r in range(hdr_row + 1, ws.max_row + 1):
            cat_val = ws.cell(r, cat_col).value
            if cat_val is None:
                continue
            if str(cat_val).strip().lower() == self.target_category:
                amt_val = ws.cell(r, amt_col).value
                try:
                    amt = float(amt_val) if amt_val not in (None, "") else 0.0
                except (ValueError, TypeError):
                    amt = 0.0
                rows.append(CollectionRow(
                    sheet=sheet_name,
                    row_num=r,
                    account_code=ws.cell(r, cat_col - 2).value,
                    customer=ws.cell(r, cat_col - 1).value,
                    category=str(cat_val).strip(),
                    details_bank=ws.cell(r, cat_col + 1).value,
                    details_company=ws.cell(r, cat_col + 2).value,
                    amount=amt,
                ))
        return rows

    def _find_header(self, ws) -> tuple[int | None, int | None]:
        """'Category' 셀이 있는 행·열 반환 (inflow 쪽 첫 번째 등장)"""
        scan_limit = min(self.header_scan_rows, ws.max_row)
        for r in range(1, scan_limit + 1):
            for c in range(1, (ws.max_column or 0) + 1):
                v = ws.cell(r, c).value
                if v and str(v).strip().lower() == "category":
                    return r, c
        return None, None

    # ──────────────────────────────────────────
    # 검증 엑셀 출력
    # ──────────────────────────────────────────

    def export_verification(
        self,
        result: CollectionResult,
        out_path: Path,
        label: str = "영업수금",
    ) -> None:
        wb_out = openpyxl.Workbook()
        ws = wb_out.active
        ws.title = "Collection_영업수금"

        # 스타일
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_font  = Font(bold=True, color="FFFFFF")
        hdr_fill  = PatternFill("solid", fgColor="1F4E79")
        total_fill = PatternFill("solid", fgColor="BDD7EE")
        total_font = Font(bold=True)

        # 제목 행
        ws.merge_cells("A1:G1")
        ws["A1"] = f"{label} — Category = Collection"
        ws["A1"].font = Font(bold=True, size=13)
        ws["A1"].alignment = Alignment(horizontal="center")

        # 헤더
        headers = [
            "Sheet (날짜)", "Account Code", "Customer",
            "Category", "Details (Bank)", "Details (Company)", "Amount",
        ]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        # 데이터
        for i, row in enumerate(result.rows, 4):
            vals = [
                row.sheet, row.account_code, row.customer,
                row.category, row.details_bank, row.details_company, row.amount,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=i, column=c, value=v)
                cell.border = border
                if c == 7:
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")

        # 합계 행
        total_row = len(result.rows) + 4
        ws.merge_cells(f"A{total_row}:F{total_row}")
        ws[f"A{total_row}"] = f"합계 (영업수금)   —   총 {len(result.rows)}건"
        ws[f"A{total_row}"].font = total_font
        ws[f"A{total_row}"].fill = total_fill
        ws[f"A{total_row}"].alignment = Alignment(horizontal="right", vertical="center")
        ws[f"A{total_row}"].border = border
        total_cell = ws.cell(total_row, 7, value=result.total)
        total_cell.number_format = "#,##0.00"
        total_cell.font = total_font
        total_cell.fill = total_fill
        total_cell.border = border

        # 열 너비
        for c, w in enumerate([28, 14, 30, 12, 45, 40, 16], 1):
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[3].height = 28

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb_out.save(out_path)
        self.logger.info(f"검증 파일 저장: {out_path}")
